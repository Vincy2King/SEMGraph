import numpy as np
import pandas as pd
# import matplotlib.pyplot as plt
from scipy import stats
from nltk.corpus import wordnet
import nltk
from openpyxl import Workbook,load_workbook # 引入头部文件
import re

def error(word):
    # r='[’!"#$%&\'()*+,-./:;<=>?@[\\]^_`{|}~]+'
    r = '[’!"#$%&\'()*+,-./:;<=>?@[\\]^_`{|}~\n。！，]+'
    line = re.sub(r, '', word)
    return line

def is_critical_wordnet(word):
    function_list = ['CC', 'IN', 'LS', 'TO', 'POS', 'RP', 'SYM', 'UH']
    is_entity_critical_word = 1
    word_list = list(word)
    pos_tags = nltk.pos_tag(word_list)[0][1]
    if pos_tags in function_list:
        is_entity_critical_word = 0
    return is_entity_critical_word


def GECO():
    wb = Workbook()
    ws = wb.active
    ws.title = 'data'

    name_list=['sentence_id','sentence','word_id','word','number_of_characters','start_with_capital_letter',
               'have_alphanumeric_letters','capital_letters_only','is_entity_critical_word',
               'number_of_dominated_nodes','complexity_score','max_dependency_distance',
               'number_of_senses_in_wordnet','avg_word_first_start_time','avg_word_first_end_time',
               'avg_word_first_duration','avg_word_go_past_time','avg_word_regression_time','avg_word_total_reading_time','pos']

    for i in range(len(name_list)):
        ws.cell(1,(i+1)).value=name_list[i]

    GECO_data = load_workbook('Gaze_data/GECO_avg.xlsx')
    GECO_data = GECO_data.get_sheet_by_name('data')
    A = GECO_data['A']
    B = GECO_data['B']
    C = GECO_data['C']
    D = GECO_data['D']
    E = GECO_data['E']
    F = GECO_data['F']
    G = GECO_data['G']
    H = GECO_data['H']
    I = GECO_data['I']
    J = GECO_data['J']
    K = GECO_data['K']
    L = GECO_data['L']
    M = GECO_data['M']
    N = GECO_data['N']
    O = GECO_data['O']
    P = GECO_data['P']
    Q = GECO_data['Q']
    R = GECO_data['R']
    S = GECO_data['S']
    T = GECO_data['T']

    real_num=2
    for i in range(1,len(A)):
        print(i/len(A))
        word=D[i].value
        new_word=error(word).strip()
        number_of_characters=len(new_word)
        wordnet_=len(wordnet.synsets(new_word))
        is_entity_critical_word=is_critical_wordnet(new_word)
        print(wordnet_)
        u=S[i].value-P[i].value

        if N[i].value==0:
            continue

        title=[A[i].value,B[i].value,C[i].value,D[i].value,number_of_characters,F[i].value,G[i].value,H[i].value,is_entity_critical_word,J[i].value,K[i].value,L[i].value,wordnet_,N[i].value,O[i].value,P[i].value,Q[i].value,R[i].value,S[i].value,T[i].value,u]
        for j in range(len(title)):
            ws.cell(real_num,(j+1)).value=title[j]
        real_num+=1

    wb.save('Gaze_data/clean_GECO_avg_4.xlsx')

GECO()
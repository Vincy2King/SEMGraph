import numpy as np
import pandas as pd
# import matplotlib.pyplot as plt
from scipy import stats

from openpyxl import Workbook,load_workbook # 引入头部文件
def eliminate_punctuate(word):
    punctuations = '"'
    if word[0]==punctuations:
        word=word[1:len(word)]
    if word[len(word)-1]==punctuations:
        word=word[:len(word)-1]
    # 显示未加标点的字符串
    return word

def draw_plot(data):    # 异常值分析
    # （1）3σ原则：如果数据服从正态分布，异常值被定义为一组测定值中与平均值的偏差超过3倍的值 → p(|x - μ| > 3σ) ≤ 0.003
    u = data.mean()  # 计算均值
    std = data.std()  # 计算标准差
    stats.kstest(data, 'norm', (u, std))
    print('均值为：%.3f，标准差为：%.3f' % (u, std))

    import matplotlib.pyplot as plt  # 导入图像库
    plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
    plt.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号

    plt.figure()  # 建立图像
    p = data.boxplot()  # 画箱线图，直接使用DataFrame的方法
    plt.show()  # 展示箱线图


def error_data(a):
    # 异常值是指样本中的个别值，其数值明显偏离其余的观测值。
    # 异常值也称离群点，异常值的分析也称为离群点的分析

    # 异常值分析 → 3σ原则 / 箱型图分析
    # 异常值处理方法 → 删除 / 修正填补

    from pandas.core.frame import DataFrame
    # a = [248,234,263,217,213,257,0,208,277]  # 列表a
    data = DataFrame(a)  # 将字典转换成为数据框
    print(data)

    draw_plot(data)

    percentile = np.percentile(a, (25, 50, 75), interpolation='midpoint')
    # 以下为箱线图的五个特征值
    Q1 = percentile[0]  # 上四分位数
    Q3 = percentile[2]  # 下四分位数
    IQR = Q3 - Q1  # 四分位距
    ulim = Q3 + 1.5 * IQR  # 上限 非异常范围内的最大值
    llim = Q1 - 1.5 * IQR  # 下限 非异常范围内的最小值
    print(ulim,llim)
    total=0
    num=0
    for b in a:
        if b>=llim and b<=ulim:
            total+=b
            num+=1
    avg=total/num
    return avg

def is_digit(arr):
    import re
    pattern = re.compile('[0-9]+')
    for v in arr:
        match = pattern.findall(v)
        if match:
            return 1
    return 0

# print(is_digit('wqefdas2sdfs'))
# a=[3030,3778,4429,3070,2748,0,0,3899,3077]
# b=[0.0, 398.0, 0.0, 395.0, 206.0, 501.0, 383.0, 841.0, 0.0]
# str_1='H_I_T123'
# print(str_1.isupper())
# error_data(b)
# eliminate_punctuate('"Yes,"')

def GECO_data_split():
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = 'data'
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = 'data'
    wb3 = Workbook()
    ws3 = wb3.active
    ws3.title = 'data'
    wb4 = Workbook()
    ws4 = wb4.active
    ws4.title = 'data'
    wb5 = Workbook()
    ws5 = wb5.active
    ws5.title = 'data'
    wb6 = Workbook()
    ws6 = wb6.active
    ws6.title = 'data'
    wb7 = Workbook()
    ws7 = wb7.active
    ws7.title = 'data'
    wb8 = Workbook()
    ws8 = wb8.active
    ws8.title = 'data'
    wb9 = Workbook()
    ws9 = wb9.active
    ws9.title = 'data'
    wb10 = Workbook()
    ws10 = wb10.active
    ws10.title = 'data'
    wb11 = Workbook()
    ws11 = wb11.active
    ws11.title = 'data'

    wb12 = Workbook()
    ws12 = wb12.active
    ws12.title = 'data'
    wb13 = Workbook()
    ws13 = wb13.active
    ws13.title = 'data'
    wb14 = Workbook()
    ws14 = wb14.active
    ws14.title = 'data'
    wb15 = Workbook()
    ws15 = wb15.active
    ws15.title = 'data'
    wb16 = Workbook()
    ws16 = wb16.active
    ws16.title = 'data'
    wb17 = Workbook()
    ws17 = wb17.active
    ws17.title = 'data'
    # wb18 = Workbook()
    # ws18 = wb18.active
    # ws18.title = 'data'
    # wb19 = Workbook()
    # ws19 = wb19.active
    # ws19.title = 'data'
    # wb20 = Workbook()
    # ws20 = wb20.active
    # ws20.title = 'data'
    # wb21 = Workbook()
    # ws21 = wb21.active
    # ws21.title = 'data'
    #
    # wb22 = Workbook()
    # ws22 = wb22.active
    # ws22.title = 'data'
    # wb23 = Workbook()
    # ws23 = wb23.active
    # ws23.title = 'data'
    # wb24 = Workbook()
    # ws24 = wb24.active
    # ws24.title = 'data'
    # wb25 = Workbook()
    # ws25 = wb25.active
    # ws25.title = 'data'
    # wb26 = Workbook()
    # ws26 = wb26.active
    # ws26.title = 'data'
    # wb27 = Workbook()
    # ws27 = wb27.active
    # ws27.title = 'data'
    # wb28 = Workbook()
    # ws28 = wb28.active
    # ws28.title = 'data'
    #
    # wb29 = Workbook()
    # ws29 = wb29.active
    # ws29.title = 'data'
    # wb30 = Workbook()
    # ws30 = wb30.active
    # ws30.title = 'data'
    # wb31 = Workbook()
    # ws31 = wb31.active
    # ws31.title = 'data'
    # wb32 = Workbook()
    # ws32 = wb32.active
    # ws32.title = 'data'
    # wb33 = Workbook()
    # ws33 = wb33.active
    # ws33.title = 'data'
    # wb34 = Workbook()
    # ws34 = wb34.active
    # ws34.title = 'data'
    # wb35 = Workbook()
    # ws35 = wb35.active
    # ws35.title = 'data'

    GECO_data = load_workbook('GECO/GECO_data_update.xlsx')
    GECO_data = GECO_data.get_sheet_by_name('data')
    A = GECO_data['A']
    B = GECO_data['B']
    C = GECO_data['C']
    D = GECO_data['D']
    E = GECO_data['E']
    F = GECO_data['F']
    G = GECO_data['G']
    H = GECO_data['H']
    I = GECO_data['I']
    J = GECO_data['J']
    K = GECO_data['K']
    L = GECO_data['L']
    M = GECO_data['M']
    N = GECO_data['N']
    O = GECO_data['O']
    P = GECO_data['P']

    for i in range(len(A)):
        print(i/len(A))
        name = [A[i], B[i], C[i], D[i], E[i], F[i], G[i], H[i], I[i], J[i], K[i], L[i], M[i], N[i], O[i], P[i]]
        if i<=16927:
            for j in range(len(name)):
                ws1.cell(row=(i+1),column=(j+1)).value=name[j].value
        elif i<=35604:
            for j in range(len(name)):
                ws2.cell(row=(i-16927),column=(j+1)).value=name[j].value
        elif i<=51306:
            for j in range(len(name)):
                ws3.cell(row=(i-35604),column=(j+1)).value=name[j].value
        elif i<=66377:
            for j in range(len(name)):
                ws4.cell(row=(i-51306),column=(j+1)).value=name[j].value
        elif i<=81031:
            for j in range(len(name)):
                ws5.cell(row=(i-66377),column=(j+1)).value=name[j].value
        elif i<=95773:
            for j in range(len(name)):
                ws6.cell(row=(i-81031),column=(j+1)).value=name[j].value
        elif i<=108750:
            for j in range(len(name)):
                ws7.cell(row=(i-95773),column=(j+1)).value=name[j].value
        elif i<=125060:
            for j in range(len(name)):
                ws8.cell(row=(i-108750),column=(j+1)).value=name[j].value
        elif i<=136294:
            for j in range(len(name)):
                ws9.cell(row=(i-125060),column=(j+1)).value=name[j].value
        elif i<=150505:
            for j in range(len(name)):
                ws10.cell(row=(i-136294),column=(j+1)).value=name[j].value
        elif i<=170179:
            for j in range(len(name)):
                ws11.cell(row=(i-150505),column=(j+1)).value=name[j].value
        elif i<=191369:
            for j in range(len(name)):
                ws12.cell(row=(i-170179),column=(j+1)).value=name[j].value
        elif i<=208356:
            for j in range(len(name)):
                ws13.cell(row=(i-191369),column=(j+1)).value=name[j].value
        elif i<=220387:
            for j in range(len(name)):
                ws14.cell(row=(i-208356),column=(j+1)).value=name[j].value
        elif i<=240911:
            for j in range(len(name)):
                ws15.cell(row=(i-220387),column=(j+1)).value=name[j].value
        elif i<=258901:
            for j in range(len(name)):
                ws16.cell(row=(i-240911),column=(j+1)).value=name[j].value
        else:
            for j in range(len(name)):
                ws17.cell(row=(i-258901),column=(j+1)).value=name[j].value
    wb1.save('GECO/split_1/GECO_data_1.xlsx')
    wb2.save('GECO/split_1/GECO_data_2.xlsx')
    wb3.save('GECO/split_1/GECO_data_3.xlsx')
    wb4.save('GECO/split_1/GECO_data_4.xlsx')
    wb5.save('GECO/split_1/GECO_data_5.xlsx')
    wb6.save('GECO/split_1/GECO_data_6.xlsx')
    wb7.save('GECO/split_1/GECO_data_7.xlsx')
    wb8.save('GECO/split_1/GECO_data_8.xlsx')
    wb9.save('GECO/split_1/GECO_data_9.xlsx')
    wb10.save('GECO/split_1/GECO_data_10.xlsx')
    wb11.save('GECO/split_1/GECO_data_11.xlsx')
    wb12.save('GECO/split_1/GECO_data_12.xlsx')
    wb13.save('GECO/split_1/GECO_data_13.xlsx')
    wb14.save('GECO/split_1/GECO_data_14.xlsx')
    wb15.save('GECO/split_1/GECO_data_15.xlsx')
    wb16.save('GECO/split_1/GECO_data_16.xlsx')
    wb17.save('GECO/split_1/GECO_data_17.xlsx')
    # wb18.save('GECO/split/GECO_data_18.xlsx')
    # wb19.save('GECO/split/GECO_data_19.xlsx')
    # wb20.save('GECO/split/GECO_data_20.xlsx')
    # wb21.save('GECO/split/GECO_data_21.xlsx')
    # wb22.save('GECO/split/GECO_data_22.xlsx')
    # wb23.save('GECO/split/GECO_data_23.xlsx')
    # wb24.save('GECO/split/GECO_data_24.xlsx')
    # wb25.save('GECO/split/GECO_data_25.xlsx')
    # wb26.save('GECO/split/GECO_data_26.xlsx')
    # wb27.save('GECO/split/GECO_data_27.xlsx')
    # wb28.save('GECO/split/GECO_data_28.xlsx')
    #
    # wb29.save('GECO/split/GECO_data_29.xlsx')
    # wb30.save('GECO/split/GECO_data_30.xlsx')
    # wb31.save('GECO/split/GECO_data_31.xlsx')
    # wb32.save('GECO/split/GECO_data_32.xlsx')
    # wb33.save('GECO/split/GECO_data_33.xlsx')
    # wb34.save('GECO/split/GECO_data_34.xlsx')
    # wb35.save('GECO/split/GECO_data_35.xlsx')


def Provo_data_split():
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = 'data'
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = 'data'
    wb3 = Workbook()
    ws3 = wb3.active
    ws3.title = 'data'
    wb4 = Workbook()
    ws4 = wb4.active
    ws4.title = 'data'
    wb5 = Workbook()
    ws5 = wb5.active
    ws5.title = 'data'
    wb6 = Workbook()
    ws6 = wb6.active
    ws6.title = 'data'
    wb7 = Workbook()
    ws7 = wb7.active
    ws7.title = 'data'
    wb8 = Workbook()
    ws8 = wb8.active
    ws8.title = 'data'
    wb9 = Workbook()
    ws9 = wb9.active
    ws9.title = 'data'
    wb10 = Workbook()
    ws10 = wb10.active
    ws10.title = 'data'
    wb11 = Workbook()
    ws11 = wb11.active
    ws11.title = 'data'

    wb12 = Workbook()
    ws12 = wb12.active
    ws12.title = 'data'
    wb13 = Workbook()
    ws13 = wb13.active
    ws13.title = 'data'
    wb14 = Workbook()
    ws14 = wb14.active
    ws14.title = 'data'

    GECO_data = load_workbook('Provo/Provo_sentence_clean_2.xlsx')
    GECO_data = GECO_data.get_sheet_by_name('data')
    A = GECO_data['A']
    B = GECO_data['B']
    C = GECO_data['C']
    D = GECO_data['D']
    E = GECO_data['E']
    F = GECO_data['F']
    G = GECO_data['G']
    H = GECO_data['H']
    I = GECO_data['I']
    J = GECO_data['J']
    K = GECO_data['K']
    L = GECO_data['L']
    M = GECO_data['M']
    N = GECO_data['N']
    O = GECO_data['O']
    P = GECO_data['P']
    Q = GECO_data['Q']

    for i in range(len(A)):
        print(i)
        name = [A[i], B[i], C[i], D[i], E[i], F[i], G[i], H[i], I[i], J[i], K[i], L[i], M[i], N[i], O[i], P[i],Q[i]]
        if i<=19907:
            for j in range(len(name)):
                ws1.cell(row=(i+1),column=(j+1)).value=name[j].value
        elif i<=37799:
            for j in range(len(name)):
                ws2.cell(row=(i-19907),column=(j+1)).value=name[j].value
        elif i<=55523:
            for j in range(len(name)):
                ws3.cell(row=(i-37799),column=(j+1)).value=name[j].value
        elif i<=72659:
            for j in range(len(name)):
                ws4.cell(row=(i-55523),column=(j+1)).value=name[j].value
        elif i<=93071:
            for j in range(len(name)):
                ws5.cell(row=(i-72659),column=(j+1)).value=name[j].value
        elif i<=117599:
            for j in range(len(name)):
                ws6.cell(row=(i-93071),column=(j+1)).value=name[j].value
        elif i<=130871:
            for j in range(len(name)):
                ws7.cell(row=(i-117599),column=(j+1)).value=name[j].value
        elif i<=148343:
            for j in range(len(name)):
                ws8.cell(row=(i-130871),column=(j+1)).value=name[j].value
        elif i<=159935:
            for j in range(len(name)):
                ws9.cell(row=(i-148343),column=(j+1)).value=name[j].value
        elif i<=172115:
            for j in range(len(name)):
                ws10.cell(row=(i-159935),column=(j+1)).value=name[j].value
        elif i <= 189587:
            for j in range(len(name)):
                ws11.cell(row=(i-172115),column=(j+1)).value=name[j].value
        elif i <= 202523:
            for j in range(len(name)):
                ws12.cell(row=(i-189587),column=(j+1)).value=name[j].value
        elif i <= 214031:
            for j in range(len(name)):
                ws13.cell(row=(i-202523),column=(j+1)).value=name[j].value
        else:
            for j in range(len(name)):
                ws14.cell(row=(i-214031),column=(j+1)).value=name[j].value
    wb1.save('Provo/split_1/Provo_data_1.xlsx')
    wb2.save('Provo/split_1/Provo_data_2.xlsx')
    wb3.save('Provo/split_1/Provo_data_3.xlsx')
    wb4.save('Provo/split_1/Provo_data_4.xlsx')
    wb5.save('Provo/split_1/Provo_data_5.xlsx')
    wb6.save('Provo/split_1/Provo_data_6.xlsx')
    wb7.save('Provo/split_1/Provo_data_7.xlsx')
    wb8.save('Provo/split_1/Provo_data_8.xlsx')
    wb9.save('Provo/split_1/Provo_data_9.xlsx')
    wb10.save('Provo/split_1/Provo_data_10.xlsx')
    wb11.save('Provo/split_1/Provo_data_11.xlsx')
    wb12.save('Provo/split_1/Provo_data_12.xlsx')
    wb13.save('Provo/split_1/Provo_data_13.xlsx')
    wb14.save('Provo/split_1/Provo_data_14.xlsx')

def combine_Provo():
    wb = Workbook()
    ws = wb.active
    ws.title = 'data'
    num=1
    for i in range(1,15):
        Provo_data = load_workbook('Provo/avg_1/Provo_avg_'+str(i)+'.xlsx')
        Provo_data = Provo_data.get_sheet_by_name('data')

        A = Provo_data['A']
        B = Provo_data['B']
        C = Provo_data['C']
        D = Provo_data['D']
        E = Provo_data['E']
        F = Provo_data['F']
        G = Provo_data['G']
        H = Provo_data['H']
        I = Provo_data['I']
        J = Provo_data['J']
        K = Provo_data['K']
        L = Provo_data['L']
        M = Provo_data['M']
        N = Provo_data['N']
        O = Provo_data['O']
        P = Provo_data['P']
        Q = Provo_data['Q']

        for j in range(1,len(A)):
            print(j)
            name = [A[j], B[j], C[j], D[j], E[j], F[j], G[j], H[j], I[j], J[j], K[j], L[j], M[j], N[j], O[j], P[j],
                    Q[j]]
            for k in range(len(name)):
                ws.cell(row=num,column=(k+1)).value=name[k].value
            num+=1
    wb.save('Provo/avg_1/Provo_avg.xlsx')

def combine_GECO():
    wb = Workbook()
    ws = wb.active
    ws.title = 'data'
    num = 1
    for i in range(1, 18):
        GECO_data = load_workbook('GECO/avg_1/GECO_avg_' + str(i) + '.xlsx')
        GECO_data = GECO_data.get_sheet_by_name('data')

        A = GECO_data['A']
        B = GECO_data['B']
        C = GECO_data['C']
        D = GECO_data['D']
        E = GECO_data['E']
        F = GECO_data['F']
        G = GECO_data['G']
        H = GECO_data['H']
        I = GECO_data['I']
        J = GECO_data['J']
        K = GECO_data['K']
        L = GECO_data['L']
        M = GECO_data['M']
        N = GECO_data['N']
        O = GECO_data['O']
        P = GECO_data['P']
        Q = GECO_data['Q']
        R = GECO_data['R']
        S = GECO_data['S']
        T = GECO_data['T']

        for j in range(1, len(A)):
            print(j)
            name = [A[j], B[j], C[j], D[j], E[j], F[j], G[j], H[j], I[j], J[j], K[j], L[j], M[j], N[j], O[j], P[j],
                    Q[j],R[j],S[j],T[j]]
            for k in range(len(name)):
                ws.cell(row=num, column=(k + 1)).value = name[k].value
            num += 1
    wb.save('GECO/avg_1/GECO_avg.xlsx')

combine_GECO()
# combine_Provo()
# GECO_data_split()
# Provo_data_split()